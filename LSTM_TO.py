from numpy import array, reshape, argmax ,arange, sum, std
import os
from openpyxl import load_workbook
from keras.models import Sequential
from keras.layers import Dense, LSTM, Dropout
import matplotlib.pyplot as plt


#코드 구조
# 1. 파일 입력 및 전처리: inputData()
# 2. 데이터 병합 및 구분(trainset, testset)처리: template()
# 3. 모델 구성 및 성능 평가: evaluate_model()
# 4. 메인 함수: __main__

INFILEPATH = "./iotlab/Data41/"	#입력 파일 경로
OUTFILEPATH = "./iotlab/"		#출력 파일 경로
NUMOFFILE = 41					#입력할 파일 수(=사람)
SHEET_SIZE = 19					#입력할 시트 수 (몇몇 파일의 20시트가 오류있음)
NUMOFTRAIN = 12					#학습에 사용할 시트 수
ACC_LIST = ["F", "G", "H"]			#l_acc_x_col, l_acc_y_col, l_acc_z_col
ACC_LIST_F = ["R", "T", "V"]		#l_acc_x_col_F, l_acc_y_col_F, l_acc_z_col_F
GYRO_LIST_F = ["L", "N", "P"]		#l_gyro_x_col_F, l_gyro_y_col_F, l_gyro_z_col_F		
LABLE_COL, TIME_COL = "B", "W"

SLICE_SIZE = 25
WINDOW_SIZE = 30		#윈도우 크기
SIZE_STEP = 3			#윈도우 내의 간격

FSELECT = {			#입력에 사용될 FEATURE
	"raw":False,				#원데이터: 6개
	"u_vector":False,			#단위벡터: 2개
	"norm_vector":True,			#벡터정규화: 6개
	"change":False,				#변화율: 6개
	"ratio":False,				#비율: 6개
	"stdDev":False,				#표준편차: 6개
	"test":False}				#테스트: 6개
NUMOFFEATURE = 6				#입력에 들어갈 feature의 개수 (FSELECT 참고)

NUMOFOUTPUT = 5			#출력 개수
EPOCH = 30				#학습 횟수

EXCEPTION = True		#불균형데이터에서 샘플이 가장 많은 레이블 제외 여부

def inputData(fPath):
	files_x = []
	files_y = []
	
	for fidx, file_name in enumerate(os.listdir(fPath)[:NUMOFFILE]):
		wb = load_workbook(filename=fPath+file_name, data_only=True)
		sheets_x = []
		sheets_y = []
		
		for sheet_idx in range(1, SHEET_SIZE+1):
			sheet_x = []
			sheet_y = []
			print("READ"+ "("+str(fidx+1)+"/"+str(NUMOFFILE)+")", file_name, "SHEET", sheet_idx)
			ws = wb[str(sheet_idx)]
			startLine = WINDOW_SIZE+2+SLICE_SIZE
			if FSELECT["stdDev"]:
				startLine += WINDOW_SIZE
			if FSELECT["ratio"]:
				startLine += SIZE_STEP
			endLine = ws.max_row-SLICE_SIZE
			
			for i in range(startLine, endLine):
				sample_x = []
				sample_y = []
				
				for winrow in range(i-WINDOW_SIZE, i+1, SIZE_STEP):
					
					for idx, char in enumerate(ACC_LIST_F+GYRO_LIST_F):
						if ws[char+str(winrow)].value == 0:
							ws[char+str(winrow)] = (ws[char+str(winrow-1)].value + ws[char+str(winrow+1)].value)/2
					
					if FSELECT["raw"]:			
						for char in ACC_LIST_F+GYRO_LIST_F:
							v = ws[char+str(winrow)].value
							sample_x.append(v)

					if FSELECT["u_vector"]:		
						normVec = []
						vSum = 0
						for idx, char in enumerate(ACC_LIST_F+GYRO_LIST_F):
							v = ws[char+str(winrow)].value
							vSum += v**2
							if idx%3 == 0:
								normVec.append(1 / vSum**(1/2))
								vSum = 0
						sample_x += normVec

					if FSELECT["norm_vector"]:		
						normVec = []
						vSum = 0
						for idx, char in enumerate(ACC_LIST_F+GYRO_LIST_F):
							v = ws[char+str(winrow)].value
							vSum += v**2
							if idx%3 == 0:
								normVec.append(1 / vSum**(1/2))
								vSum = 0
						for idx, char in enumerate(ACC_LIST_F+GYRO_LIST_F):
							v = ws[char+str(winrow)].value
							if idx < 3:
								sample_x.append(v*normVec[0])
							else:
								sample_x.append(v*normVec[1])
					
					if FSELECT["change"]:	
						for idx, char in enumerate(ACC_LIST_F+GYRO_LIST_F):
							v2 = ws[char+str(winrow)].value
							v1 = ws[char+str(winrow-SIZE_STEP)].value
							v = (v2 - v1) / v1
							sample_x.append(v)
					
					if FSELECT["ratio"]:		
						rlist = []
						for idx, char in enumerate(ACC_LIST_F+GYRO_LIST_F):
							v = ws[char+str(winrow)].value
							rlist.append(v)
						s = 0
						for idx, v in enumerate(rlist):
							if idx%3 == 0:
								s = sum(rlist[idx:idx+3])
							sample_x.append(v/s)
					
					if FSELECT["stdDev"]:
						for idx, char in enumerate(ACC_LIST_F+GYRO_LIST_F):
							dlist = []
							for drow in range(winrow, winrow-WINDOW_SIZE, -SIZE_STEP):
								dv = ws[char+str(drow)].value
								dlist.append(dv)
							v = std(dlist)
							sample_x.append(v)
					
					if FSELECT["test"]:	
						for i in range(NUMOFFEATURE):
							sample_x.append(v)

			
				sheet_x.append(sample_x)
				
		
				lable_value = ws[LABLE_COL+str(i)].value
				for i in range(int(NUMOFOUTPUT)):
					sample_y.append(0)
				sample_y[lable_value] = 1
				sheet_y.append(sample_y)
			
		
			sheets_x.append(sheet_x)
			sheets_y.append(sheet_y)
		
		files_x.append(sheets_x)
		files_y.append(sheets_y)
	return files_x, files_y



def template(files, sheet_size, nTrain, dType="int"):
	sample_train = []
	sample_test = []
	for person_idx, person in enumerate(files):
		print("TEMPLATE", str(person_idx+1)+"/"+ str(NUMOFFILE))
		for sheet_idx, sheet in enumerate(person):
			for line in sheet:
				
				if sheet_idx%sheet_size < nTrain:
					sample_train.append(line)
				else:
					sample_test.append(line)

	
	if dType == "float":
		sample_train = reshape(sample_train, (array(sample_train).shape[0], int(array(sample_train).shape[1]/NUMOFFEATURE), NUMOFFEATURE))
		sample_test = reshape(sample_test, (array(sample_test).shape[0], int(array(sample_test).shape[1]/NUMOFFEATURE), NUMOFFEATURE))
	return array(sample_train), array(sample_test)



def evaluate_model(trainX, trainY, testX, testY):
	
	model = Sequential()
	model.add(LSTM(128, input_shape=(None, NUMOFFEATURE), return_sequences=True))
	model.add(LSTM(128, return_sequences=False))
	model.add(Dropout(0.2))
	model.add(Dense(NUMOFOUTPUT, activation='relu'))
	model.add(Dense(NUMOFOUTPUT, activation='softmax'))
	model.compile(loss='categorical_crossentropy', optimizer='sgd', metrics=['accuracy'])
	model.summary()


	history = model.fit(trainX, trainY, batch_size=32, epochs=EPOCH, validation_data=(testX, testY))
	

	y_vloss = history.history['val_loss']
	y_loss = history.history['loss']
	x_len = arange(len(y_loss))
	plt.plot(x_len, y_vloss, marker='.', c="red", label='Testset_loss')
	plt.plot(x_len, y_loss, marker='.', c="blue", label='Trainset_loss')
	plt.legend(loc='upper right')
	plt.grid()
	plt.xlabel('epoch')
	plt.ylabel('loss')
	plt.show()

	return model
	
 

if __name__ == "__main__":

	files_x, files_y = inputData(INFILEPATH)
	

	x_train, x_test = template(files_x, SHEET_SIZE, NUMOFTRAIN, dType="float")
	y_train, y_test = template(files_y, SHEET_SIZE, NUMOFTRAIN, dType="int")


	print("TRAINING")
	print(array(x_train).shape, array(y_train).shape)
	model = evaluate_model(x_train, y_train, x_test, y_test)


	predict = model.predict(x_test, batch_size=32)
	numOfCase, posSum = 0, 0
	visX, visY = [], []
	with open(OUTFILEPATH+"result.csv","w") as f:
		for idx, line in enumerate(predict):
			print("WRITE PROCESS", str(idx)+"/"+str(len(predict)))
			p1 = predict[idx]
			y1 = y_test[idx]
			p2 = None
			y2 = argmax(y1)
			
			if EXCEPTION:
				p2 = argmax(p1[1:])+1
			else:	
				p2 = argmax(p1)

			visX.append(p2)
			visY.append(y2)

			if y2 != 0:
				numOfCase += 1
				if p2 == y2:
					posSum += 1

			f.write(",".join([str(x) for x in p1])+",")
			f.write(",".join([str(x) for x in y1])+",")
			f.write(str(p2)+",")
			f.write(str(y2)+",")
			f.write("\n")
	plt.xlabel("sample(20ms/sample)")
	plt.ylabel("class")
	plt.plot(visX, 'x', label='predict')
	plt.plot(visY, '-', label='r_value')
	plt.yticks([0,1,2,3,4], ["NONE","L HS","L TO","R HS","R TO"])
	plt.legend()
	plt.show()
	print("ACCURACY", posSum/numOfCase*100, "CASE", numOfCase)

		

